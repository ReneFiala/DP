import pandas as pd
import scipy.stats as stats
from pathlib import Path
import os
import statsmodels.api as sm
from statsmodels.formula.api import ols
import numpy as np
import openpyxl as xl

ROOT = Path("./2025-08-18_test")
METHOD_NAMES = ['rnd_0', 'rnd_1', 'rnd_2', 'rnd_3',
           'sp_ag', 'sp_ag_num-c', 'sp_ag_num-w',
           'oi-3b', 'oi-3b-qrt', 'oi-5b',
           'oi-3b_sp', 'oi-3b-qrt_sp', 'oi-5b_sp',
           'lai-3b', 'lai-5b', 'lai-3b_sp', 'lai-5b_sp',
           'pv-3b', 'pv-3b_sp', 'pv-5b', 'pv-5b_sp']
           # 'kmeans_all']
PRINT_RESULTS = True
TESTS_ALPHA = 0.01 # hladina významnosti

#%% Load data into dict{ method_name: list[list[float]] }

def get_method_names(root: Path) -> list[str]:
    dirs = os.listdir(root)
    methods = set([x[:-6] for x in dirs])
    times = [os.path.getmtime(root / f"{x}-t0-v1" / "log_test.csv") for x in methods]
    # return sorted(methods)
    return [x for _, x in sorted(zip(times, methods))]

def load_method(name: str, metric: str) -> list:
    output = []
    for t in range(5):
        suboutput = []
        for v in range(5):
            if t == v:
                continue
            # df = pd.read_csv(ROOT / f"{name}-t{t}-v{v}" / "aver_counts.csv", sep=";")
            df = pd.read_csv(ROOT / f"{name}-t{t}-v{v}" / "log_test.csv", sep=";")
            suboutput.append(df[metric][0])
        output.append(suboutput)
    return output

def load_all_methods(methods: list[str], metric: str) -> dict:
    metrics = [load_method(x, metric) for x in methods]
    return {k: v for k, v in zip(methods, metrics)}

#%% Utilities

GREEN = "\033[32m"
RED = "\033[31m"
RESET = "\033[0m"

def flatten_list(xss):
    return [x for xs in xss for x in xs]

def do_tests(test, data, flatten=False, unpack=False, kw={}, attr=None):
    key_length = get_key_length(data)
    results = []
    for k, v in data.items():
        if flatten:
            v = flatten_list(v)
        if unpack:
            result = test(*v, **kw)
        else:
            result = test(v)
        if attr:
            result = result._asdict()[attr]
        results.append(result)
        if PRINT_RESULTS:
            if attr == "pvalue":
                print_pvalue(k, result, key_length)
            else:
                print(f"{k.ljust(key_length)}\t{result:.4f}")
    return results

def cov(v):
    return np.std(v) / np.average(v)

def get_key_length(data):
    return max(len(x) for x in data.keys())

def print_pvalue(k, result, key_length):
    print(f"{k.ljust(key_length)}\t{result:.4f}", end="")
    if result < TESTS_ALPHA:
        print(RED + " Rejected")
    else:
        print(GREEN + " OK")
    print(RESET, end="")
    
def print_table(data) -> None:
    def print_hr(lengths) -> None:
        print("-" * (sum(lengths) + 3*len(lengths) + 1))
    lengths = [0 for x in data[0]]
    for row in data:
        i = 0
        for cell in row:
            lengths[i] = max(lengths[i], len(str(cell)))
            i += 1
    i = 0
    print_hr(lengths)
    for row in data:
        for cell, l in zip(row, lengths):
            print("| " + str(cell).rjust(l), end=" ")
        print("|")
        if i == 0:
            print_hr(lengths)
        i += 1
    print_hr(lengths)

def get_statsmodels_anova(data):
    anovas = []
    groups = []
    for t in range(5):
        for v in range(4):
            groups.append(f"T{t}")
    for k, v in data.items():
        values = flatten_list(v)
        df = pd.DataFrame({"Metric": values, "Fold": groups})
        model = ols('Metric ~ Fold', data=df).fit()
        anovas.append(sm.stats.anova_lm(model, typ=2))
    return anovas

def sm_get_anova_p(data, anovas):
    results = []
    key_length = get_key_length(data)
    for m, a in zip(data.keys(), anovas):
        p = a["PR(>F)"]['Fold']
        if PRINT_RESULTS:
            print_pvalue(m, p, key_length)
        results.append(p)
    return results

def sm_get_anova_eta(data, anovas):
    results = []
    key_length = get_key_length(data)
    for m, a in zip(data.keys(), anovas):
        eta = a['sum_sq']['Fold'] / (a['sum_sq']['Fold'] + a['sum_sq']['Residual'])
        if PRINT_RESULTS:
            print(f"{m.ljust(key_length)}\t{eta:.4f}")
        results.append(eta)
    return results
        
def sm_get_anova_omega(data, anovas):
    results = []
    key_length = get_key_length(data)
    for m, a in zip(data.keys(), anovas):
        mse = a['sum_sq']['Residual'] / a['df']['Residual']
        omega = (a['sum_sq']['Fold'] - a['df']['Fold'] * mse) / (a['sum_sq']['Fold'] + a['sum_sq']['Residual'] + mse)
        if PRINT_RESULTS:
            print(f"{m.ljust(key_length)}\t{omega:.4f}")
        results.append(omega)
    return results

def print_tukey(data, k, result):
    key_length = get_key_length(data)
    shape = result.pvalue.shape
    out_matrix = [ [""] * (shape[0] + 1) for i in range(shape[0] + 1)]
    for i in range(shape[0]):
        cell = f"T{i+1}"
        out_matrix[0][i+1] = cell
        out_matrix[i+1][0] = cell
    out_matrix[0][0] = k.ljust(key_length)
    for i in range(shape[0]):
        for j in range(shape[0]):
            if i == j:
                out_matrix[i+1][j+1] = "---"
            else:
                out_matrix[i+1][j+1] = f"{result.pvalue[i][j]:.4f}"
    print_table(out_matrix)
    
def print_tukey_simple(data, k, result):
    shape = result.pvalue.shape
    out_matrix = [ [""] * (shape[0] + 1) for i in range(shape[0] + 1)]
    for i in range(shape[0]):
        cell = f"T{i+1}"
        out_matrix[0][i+1] = cell
        out_matrix[i+1][0] = cell
    out_matrix[0][0] = k
    for i in range(shape[0]):
        for j in range(shape[0]):
            if i == j:
                out_matrix[i+1][j+1] = "---"
            else:
                out_matrix[i+1][j+1] = f"{result.pvalue[i][j]:.4f}"
    for row in out_matrix:
        print("\t".join(row).replace(".", ","))

def do_tukey(data):
    results = []
    for k, v in data.items():
        result = stats.tukey_hsd(*v)
        results.append(result.pvalue)
        if PRINT_RESULTS:
            print_tukey(data, k, result)
    return results


def do_all_for_metric(metric, simple_names=False):
    data = load_all_methods(METHOD_NAMES, metric)
    # data = {"rnd_a": [[0.744121715,0.71875,0.755244755,0.737609329],
    # [0.793478261,0.823747681,0.777777778,0.816696915],
    # [0.723004695,0.709677419,0.726717557,0.741433022],
    # [0.80245023,0.811463047,0.788161994,0.768532526],
    # [0.762857143,0.771186441,0.758915835,0.755304102]]}
    
    df = pd.DataFrame({"method" if simple_names else "Split\nmethod": data.keys()})
    
    if PRINT_RESULTS:
            print("\nStep 1: Normality test / Shapiro-Wilk")
    res = do_tests(stats.shapiro, data, flatten=True, attr="pvalue")
    df['shapiro-wilk' if simple_names else 'Shapiro-Wilk\np-val'] = res
    
    if PRINT_RESULTS:
            print("\nStep 2: Homogeneity of variances test / Levene")
    res = do_tests(stats.levene, data, unpack=True, kw={"center": "mean"}, attr="pvalue")
    df['levene' if simple_names else 'Levene\np-val'] = res
    
    if PRINT_RESULTS:
            print("\nStep 2.5: Homogeneity of variances test / Bartlett")
    res = do_tests(stats.bartlett, data, unpack=True, attr="pvalue")
    # df['Bartlett\np-val'] = res
    
    if PRINT_RESULTS:
            print("\nStep 3: Homogeneity of means test / One-way ANOVA")
    # do_tests(stats.f_oneway, data, unpack=True)
    anovas = get_statsmodels_anova(data)
    res = sm_get_anova_p(data, anovas)
    df['anova' if simple_names else 'ANOVA\np-val'] = res
    
    if PRINT_RESULTS:
            print("\nStep 4: Effect sizes")
            print("Eta-square")
    res = sm_get_anova_eta(data, anovas)
    df['eta_squared' if simple_names else 'η²'] = res
    if PRINT_RESULTS:
            print("Omega-square")
    res = sm_get_anova_omega(data, anovas)
    df['omega_squared' if simple_names else 'ω²'] = res
    
    if PRINT_RESULTS:
            print("\nStep 5: Fold-fold homogeneity of means / Tukey HSD")
    tukey = do_tukey(data)
    n_bad_pairs = [np.sum(np.where(x < TESTS_ALPHA, True, False))
                   for x in tukey]
    df['tukey_pairs' if simple_names else 'Tukey\np<α pairs'] = n_bad_pairs
    
    if PRINT_RESULTS:
            print("\nStep 6: Coefficient of variability for full CV")
    res = do_tests(cov, data, flatten=True)
    df['var_coeff' if simple_names else 'Variation\nCoeff.'] = res
    
    res = do_tests(np.average, data, flatten=True)
    df['average' if simple_names else 'Average\nmetric val.'] = res
    
    res = do_tests(np.std, data, flatten=True)
    df['stdev' if simple_names else 'Standard\ndeviation'] = res
    return {
        "tests": df,
        "tukey": tukey,
    }

# deprecated
def metrics_and_save(metric):
    df, tukey, method_names = do_all_for_metric(metric)
    df.to_csv(f"./stats_out/stats_{metric}_main.csv", index=False, sep=";", decimal=",")
    for k, v in zip(method_names, tukey):
        df_t = pd.DataFrame(v)
        df_t.to_csv(f"./stats_out/stats_{metric}_tukey_{k}.csv", index=False, sep=";", decimal=",", header=False)

def rect_border(ws, x1, x2, y1, y2, inner_style, outer_style):
    for x in range(x1, x2+1):
        for y in range(y1, y2+1):
            border = xl.styles.Border(
                left=xl.styles.Side(border_style=outer_style if x == x1 else inner_style),
                right=xl.styles.Side(border_style=outer_style if x == x2 else inner_style),
                top=xl.styles.Side(border_style=outer_style if y == y1 else inner_style),
                bottom=xl.styles.Side(border_style=outer_style if y == y2 else inner_style),
            )
            cell = ws.cell(row=y, column=x)
            cell.border = border

def calc_and_save(metrics, out_name):
    wb = xl.Workbook()
    wb.remove(wb.active)
    for metric in metrics:
        results = do_all_for_metric(metric)
        add_sheet(wb, metric, results["tests"], results["tukey"])
    wb.save(out_name)
    return results;

def calc_and_save_csv(metrics, out_name):
    for metric in metrics:
        results = do_all_for_metric(metric, True)
        results['tests'].to_csv(f"{out_name}_{metric}.csv", index=False)
    return results;

def add_sheet(workbook, title, df, tukey):
    ws = workbook.create_sheet(title)
    ws.title = title
    ws.sheet_view.zoomScale = 130
    ws.column_dimensions['A'].width = 13.4
    
    data = []
    data.append(list(df.columns))
    for _, row in df.iterrows():
        data.append(list(row))
        
    for row in data:
        ws.append(row)
    
    rows = len(data)
    for r in range(rows):
        cols = len(data[r])
        r += 1
        for c in range(cols):
            c += 1
            cell = ws.cell(row=r, column=c)
            if c == 1 and r > 1:
                cell.font = xl.styles.Font(italic=True)
            if r == 1:
                cell.font = xl.styles.Font(bold=True)
                cell.alignment = xl.styles.Alignment(wrapText=True)
            if c > 1 and r > 1:
                cell.number_format = "0.0000"
            if c == 7 and r > 1:
                cell.number_format = "0"
            if c == 8 and r > 1:
                cell.number_format = "0.00%"
    rect_border(ws, 1, len(data[0]), 1, len(data), 'thin', 'medium')
    rect_border(ws, 1, len(data[0]), 1, 1, 'thin', 'medium')
    rect_border(ws, 1, 1, 1, len(data), 'thin', 'medium')
    rect_border(ws, 1, 1, 1, 1, 'thin', 'medium')
    
    red_fill = xl.styles.PatternFill(patternType="solid", bgColor="FFAAAA")
    ws.conditional_formatting.add(f'B2:E{rows}', xl.formatting.rule.CellIsRule(
        operator="lessThan", formula=[f'{TESTS_ALPHA}'], fill=red_fill)
    )
    
    
    for col, start in zip(['E', 'F', 'G', 'H', 'J'], ['num', 'num', 'num', 'percentile', 'percentile']):
        ws.conditional_formatting.add(
            f'{col}2:{col}{len(data)}',
            xl.formatting.rule.ColorScaleRule(
                start_type=start, start_value=0, start_color='AAFFAA',
                mid_type='percentile', mid_value=50, mid_color='FFFFAA',
                end_type='percentile', end_value=100, end_color='FFAAAA'
            )
        )
    
    for col, start in zip(['I'], ['percentile']):
        ws.conditional_formatting.add(
            f'{col}2:{col}{len(data)}',
            xl.formatting.rule.ColorScaleRule(
                start_type=start, start_value=0, start_color='FFAAAA',
                mid_type='percentile', mid_value=50, mid_color='FFFFAA',
                end_type='percentile', end_value=100, end_color='AAFFAA'
            )
        )
        
    last_row = len(data)+2
    cell = ws.cell(row=last_row, column=1)
    cell.value = "Tukey pair tests"
    cell.font = xl.styles.Font(bold=True)
    last_row += 1
    
    for t, m in zip(tukey, data[1:]):
        app = []
        header = [m[0]]
        header.extend([f"T{x}" for x in range(1, 6)])
        app.append(header)
        for i, tr in enumerate(t):
            line = [f"T{i+1}"]
            line.extend(tr)
            line[i+1] = "---"
            app.append(line)
        for line in app:
            ws.append(line)
        rect_border(ws, 1, len(line), last_row, last_row+len(app)-1, 'thin', 'medium')
        ws.cell(row=last_row, column=1).font = xl.styles.Font(bold=True)
        for x in range(2, len(line)+1):
                ws.cell(row=last_row, column=x).font = xl.styles.Font(italic=True)
        for y in range(2, len(app)):
                ws.cell(row=last_row+y, column=1).font = xl.styles.Font(italic=True)
        ws.conditional_formatting.add(f'B{last_row}:F{last_row+len(app)-1}', xl.formatting.rule.CellIsRule(
            operator="lessThan", formula=[f'{TESTS_ALPHA}'], fill=red_fill)
        )
        for x in range(2, len(line)+1):
            for y in range(1, len(app)):
                ws.cell(row=last_row+y, column=x).number_format = "0.0000"
    
        last_row += len(app)

# result = calc_and_save(["ap_sum_0.3"], "./stats_pyxl_4.xlsx")
# calc_and_save(["ap_sum_0.3", "ap_avg_0.3", "ap_sum_0.5", "ap_avg_0.5", "loss"], "./detailed_results/11-18-2/main.xlsx")
# calc_and_save(["armse", "amae", "arsq"], "./detailed_results/11-18-2/avers.xlsx")
# calc_and_save_csv(["ap_sum_0.3", "ap_avg_0.3", "ap_sum_0.5", "ap_avg_0.5", "loss"], "./detailed_results/11-18-2/")
# calc_and_save(["rsq", "rmse", "mae"], "./stats_count.xlsx")
# calc_and_save(["rsq"], "./detailed_results/prezentace_rsq.xlsx")
# calc_and_save(["arsq"], "./detailed_results/prezentace_counts.xlsx")
calc_and_save(["ap_sum_0.3", "ap_avg_0.3", "fscore_sum_0.3", "fscore_avg_0.3",
               "ap_sum_0.5", "ap_avg_0.5", "fscore_sum_0.5", "fscore_avg_0.5"], "./detailed_results/prezentace.xlsx")

#%%

# for k, v in zip(x['tests']['Split\nmethod'], x['tukey']):
#     np.savetxt(f"./tukey_{k}.csv", v, fmt="%.4f", delimiter=",")